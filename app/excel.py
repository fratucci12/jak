from pathlib import Path


def gerar_excel(propostas, filename):
    from loguru import logger as log
    import math
    import unicodedata
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    log.info(
        "Início da geração do Excel com template. Recebido tipo: %s, quantidade: %s",
        type(propostas),
        len(propostas),
    )

    template_path = Path(__file__).resolve().parent / "Classificação.xlsx"
    wb = load_workbook(template_path)
    ws = wb["Plan1"]

    template_last_row = ws.max_row
    style_refs = [ws.cell(row=template_last_row, column=col)._style for col in range(1, 7)]

    for row in ws.iter_rows(min_row=2, max_row=template_last_row):
        for cell in row:
            cell.value = None

    def _valor_unitario(item):
        valores = item.get("valores") or {}
        bloco = valores.get("valorPropostaInicialOuLances") or {}
        valor = bloco.get("valorInformado")
        if isinstance(valor, dict):
            return valor.get("valorUnitario")
        return valor

    def _quantidade(item):
        return (
            item.get("quantidadeOfertada")
            or item.get("quantidadeSolicitada")
            or item.get("quantidade")
        )

    def _valor_total(item, valor_unitario, quantidade):
        valores = item.get("valores") or {}
        bloco = valores.get("valorPropostaInicialOuLances") or {}
        calculado = bloco.get("valorCalculado") or {}
        total = calculado.get("valorTotal")
        if total is None:
            bloco = valores.get("valorPropostaInicial") or {}
            calculado = bloco.get("valorCalculado") or {}
            total = calculado.get("valorTotal")
        if total is None and valor_unitario is not None and quantidade:
            total = valor_unitario * quantidade
        return total

    propostas_ordenadas = sorted(
        propostas,
        key=lambda p: p.get("classificacao") or math.inf,
    )

    def _clean_text(value: str | None):
        if not value:
            return None
        cleaned = "".join(
            ch for ch in value if not unicodedata.category(ch).startswith("C")
        )
        cleaned = " ".join(cleaned.split())
        return cleaned or None

    for idx, p in enumerate(propostas_ordenadas, start=2):
        fornecedor = (p.get("participante") or {}).get("nome")
        if fornecedor and len(fornecedor) > 40:
            fornecedor = fornecedor[:37].rstrip() + "..."
        marca = p.get("marcaFabricante") or ""
        modelo = p.get("modeloVersao") or ""
        marca_modelo = (
            _clean_text(f"{marca}/{modelo}".strip("/")) if (marca or modelo) else None
        )
        valor_unitario = _valor_unitario(p)
        quantidade = _quantidade(p)
        valor_total = _valor_total(p, valor_unitario, quantidade)

        if idx > template_last_row:
            for col in range(1, 7):
                ws.cell(row=idx, column=col)._style = style_refs[col - 1]

        ws.cell(row=idx, column=1, value=p.get("classificacao") or (idx - 1))
        ws.cell(row=idx, column=2, value=fornecedor)
        ws.cell(row=idx, column=3, value=marca_modelo)
        ws.cell(row=idx, column=4, value=valor_unitario)
        ws.cell(row=idx, column=5, value=quantidade)
        ws.cell(
            row=idx,
            column=6,
            value=valor_total if valor_total is not None else f"=E{idx}*D{idx}",
        )

        log.info(
            "Linha %s preenchida: fornecedor=%s valor=%s qtd=%s total=%s",
            idx,
            fornecedor,
            valor_unitario,
            quantidade,
            valor_total,
        )

    for col in range(1, 7):
        letter = get_column_letter(col)
        max_length = 0
        for cell in ws[letter]:
            if cell.value is not None:
                value = str(cell.value).strip()
                max_length = max(max_length, len(value))
        padding = 0 if col == 3 else (8 if col == 6 else 2)
        ws.column_dimensions[letter].width = max_length + padding

    wb.save(filename)
    log.info("Excel salvo em %s", filename)
    return filename
